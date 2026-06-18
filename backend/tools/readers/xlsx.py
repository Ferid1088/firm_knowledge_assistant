"""XLSXReaderTool — extract tables and cell data from Excel files."""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from backend.core.tool_base import ToolMetadata
from backend.tools.readers.base import FileReaderTool, RawContent

logger = logging.getLogger(__name__)


class XLSXReaderTool(FileReaderTool):
    """Extract one table per sheet from Excel workbooks via openpyxl."""

    metadata = ToolMetadata(
        name="reader:xlsx",
        version="1.0.0",
        tool_type="reader",
        description="Extract tables and cell data from Excel files",
        dependencies=["openpyxl>=3.0", "pandas>=1.0"],
        capabilities={
            "formats": [".xlsx", ".xls", ".XLSX"],
            "max_size": 50 * 1024 * 1024,
            "supports_tables": True,
            "supports_multiple_sheets": True,
        },
        is_production=True,
        is_async=True,
    )

    supported_extensions = [".xlsx", ".xls", ".XLSX"]
    format_name = "xlsx"

    async def execute(self, input_data: str, **kwargs) -> RawContent:
        """Read the workbook with data_only=True and return all sheets as table dicts."""
        file_path = str(input_data)
        is_valid, error = await self.validate_input(file_path)
        if not is_valid:
            return RawContent(
                format="xlsx",
                content="",
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                is_corrupted=True,
                warnings=[error],
            )

        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, data_only=True)
            tables_data = []
            text_parts = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_data = []
                for row in ws.iter_rows(values_only=True):
                    str_row = [str(c) if c is not None else "" for c in row]
                    rows_data.append(str_row)

                tables_data.append({
                    "sheet": sheet_name,
                    "rows": len(rows_data),
                    "cols": max((len(r) for r in rows_data), default=0),
                    "data": rows_data,
                })
                text_parts.append(f"=== Sheet: {sheet_name} ===")
                for row in rows_data:
                    text_parts.append(" | ".join(row))

            p = Path(file_path)
            return RawContent(
                format="xlsx",
                content="\n".join(text_parts),
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                table_count=len(tables_data),
                tables=tables_data,
                file_size=p.stat().st_size,
                modified_time=datetime.fromtimestamp(p.stat().st_mtime).isoformat(),
            )

        except Exception as e:
            return self._handle_error(e, file_path)
